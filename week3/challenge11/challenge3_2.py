#!/usr/bin/env python

from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
import sys

def combine(srcfile):
    '''
    该函数可以处理原数据文件:
    1. 合并表格并写入的 combine 表中
    2. 保存原数据文件
    '''
    wb = load_workbook(filename=srcfile)
    sheet_students = wb['students']
    sheet_time = wb['time']

    students_dict = {}

    # 获取列名索引
    spamiter = sheet_students.iter_rows()
    head_row = next(spamiter)
    header = [x.value for x in head_row]
    create_time = header.index('创建时间')
    course_name = header.index('课程名称')
    students = header.index('学习人数')

    for row in sheet_students.iter_rows(min_row=2):
        students_dict[row[course_name].value] = [row[create_time].value,row[students].value]


    spamiter = sheet_time.iter_rows()
    head_row = next(spamiter)
    header = [x.value for x in head_row]
    course_name = header.index('课程名称')
    learn_time = header.index('学习时间')


    for row in sheet_time.iter_rows(min_row=2):
        students_dict[row[course_name].value].append(row[learn_time].value)
        #print(row[learn_time].value)

    # 写入 combine 表中
    #if not wb['combine']:
    #else:
    #    sheet_combine = wb['combine']
    #    wb.remove(sheet_combine)
    #    sheet_combine = wb.create_sheet(title='combine')
    try:
        sheet_combine = wb['combine']
        wb.remove(sheet_combine)
    except:
        pass
    sheet_combine = wb.create_sheet(title='combine')
    sheet_combine.append(['创建时间','课程名称','学习人数','学习时间'])
    for k in students_dict.keys():
        t = [students_dict[k][0], k, students_dict[k][1], students_dict[k][2]]
        sheet_combine.append(t)

    wb.save(filename=srcfile)


def split(srcfile):
    '''
    该函数可以分割文件:
    1. 读取 combine 表中的数据
    2. 写入不同的数据表中
    '''
    data_dict = {}
    wb_course = load_workbook(filename=srcfile)
    sheet_combine = wb_course['combine']

    for row in sheet_combine.iter_rows(min_row=2):
        t = [x.value for x in row]
        data_dict.setdefault(row[0].value.year,[]).append(t)

    for y in data_dict:
        wb = Workbook()
        ws = wb.active
        ws.title = 'combine'
        ws.append(['创建时间','课程名称','学习人数','学习时间'])
        for t in data_dict[y]:
            ws.append(t)
        else:
            wb.save(filename='{:d}.xlsx'.format(y))





# 执行
if __name__ == '__main__':
    try:
        srcfile = sys.argv[1]
    except IndexError:
        print('Parameter Error')
        sys.exit(1)
    combine(srcfile)
    split(srcfile)
